import re
import math
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from config import EXCESSIVE_MONTHLY_RAINFALL_MM


def _instrument_sort_key(row: dict) -> int:
    match = re.search(r"Instrument-(\d+)", row.get("station", ""), re.IGNORECASE)
    return int(match.group(1)) if match else float("inf")


# Hex fills for cell status colors
_FILLS = {
    "green":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "red":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "grey":   PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "orange": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
    "header": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
}

# Human-readable header overrides (columns not listed fall back to title-cased snake_case)
_HEADER_NAMES = {
    "station":               "Station",
    "expected_obs":          "Expected Obs",
    "actual_obs":            "Actual Obs",
    "overall_uptime":        "Overall Uptime",
    "rain_gauge_1_total":    "Rain Gauge 1 Total (mm)",
    "rain_gauge_2_total":    "Rain Gauge 2 Total (mm)",
    "rainfall_pct_diff":     "Rainfall % Diff",
    "rain_gauge_corr":       "Rain Gauge Correlation",
    "rain_gauge_max_diff":   "Max Daily Diff (mm)",
    "rain_gauge_max_diff_day": "Max Diff Date",
    "rain_gauge_1_zero_days": "Rain Gauge 1 Zero-Rain Days",
    "rain_gauge_2_zero_days": "Rain Gauge 2 Zero-Rain Days",
}

# Which columns get color-coded and what their status column is named
# Format: { data_column: status_column }
# The status column is expected to hold 'green' | 'yellow' | 'red'
_COLOR_MAP = {
    "overall_uptime": "uptime_status",
}


def generate_report(summary_rows: list[dict], output_path: Path) -> None:
    """
    Write a color-formatted Excel workbook from a list of per-station summary dicts.

    Each dict should be the output of stats.summarize_station().
    Status columns (e.g. 'uptime_status') drive cell color; they are hidden
    from the sheet after coloring.
    """
    summary_rows = sorted(summary_rows, key=_instrument_sort_key)
    df = pd.DataFrame(summary_rows)

    # Columns to hide after they've been used for coloring
    status_cols = set(_COLOR_MAP.values())
    display_cols = [c for c in df.columns if c not in status_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Station Health"

    header_font = Font(bold=True, color="FFFFFF")

    # Write header row
    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADER_NAMES.get(col_name, col_name.replace("_", " ").title()))
        cell.fill = _FILLS["header"]
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data rows (convert NaN → None so cells are blank, not "nan")
    for row_idx, row in enumerate(df[display_cols].itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and math.isnan(value):
                value = "N/A"
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply color coding
    for data_col, status_col in _COLOR_MAP.items():
        if data_col not in display_cols or status_col not in df.columns:
            continue
        col_idx = display_cols.index(data_col) + 1
        for row_idx, status in enumerate(df[status_col], start=2):
            fill = _FILLS.get(status)
            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Yellow highlight: one gauge has N/A while the other has a real value
    rg_pair = ("rain_gauge_1_total", "rain_gauge_2_total")
    if all(c in display_cols for c in rg_pair):
        idx1 = display_cols.index(rg_pair[0]) + 1
        idx2 = display_cols.index(rg_pair[1]) + 1
        for row_idx in range(2, len(df) + 2):
            v1 = ws.cell(row=row_idx, column=idx1).value
            v2 = ws.cell(row=row_idx, column=idx2).value
            v1_na = v1 == "N/A"
            v2_na = v2 == "N/A"
            if v1_na and not v2_na and v2 is not None:
                ws.cell(row=row_idx, column=idx1).fill = _FILLS["yellow"]
            elif v2_na and not v1_na and v1 is not None:
                ws.cell(row=row_idx, column=idx2).fill = _FILLS["yellow"]

    # Red highlight: Rain Gauge total > 1000 mm (likely sensor error)
    for rg_col in ("rain_gauge_1_total", "rain_gauge_2_total"):
        if rg_col in display_cols:
            col_idx = display_cols.index(rg_col) + 1
            for row_idx in range(2, len(df) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)) and val > EXCESSIVE_MONTHLY_RAINFALL_MM:
                    ws.cell(row=row_idx, column=col_idx).fill = _FILLS["red"]

    # Orange highlight: Rainfall % Diff > 100%
    if "rainfall_pct_diff" in display_cols:
        pct_col_idx = display_cols.index("rainfall_pct_diff") + 1
        for row_idx in range(2, len(df) + 2):
            val = ws.cell(row=row_idx, column=pct_col_idx).value
            if isinstance(val, (int, float)) and val > 100:
                ws.cell(row=row_idx, column=pct_col_idx).fill = _FILLS["orange"]

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_path)
    print(f"[report] Saved: {output_path}")
